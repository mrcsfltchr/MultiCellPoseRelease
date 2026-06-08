"""Convert QuantGUV Excel exports to labelled instance-mask arrays.

QuantGUV reports accepted GUVs in an Excel sheet. The current upstream GUI
stores keypoint centers internally, but older exports may omit them. This
adapter requires center columns (``x``/``y`` or ``center_x``/``center_y``) and
uses either diameter or area to render filled circular instance masks.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
from openpyxl import load_workbook
from PIL import Image


IMG_EXTS = (".tif", ".tiff", ".png", ".jpg", ".jpeg", ".bmp")


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", required=True, help="QuantGUV exported .xlsx file.")
    parser.add_argument("--image-dir", required=True, help="Directory containing original images.")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--sheet", default=None, help="Excel sheet name. Defaults to active sheet.")
    parser.add_argument("--scale-factor", type=float, default=None, help="Microns per pixel, if diameter is in microns.")
    parser.add_argument("--recursive", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args(argv)


def norm_col(name: object) -> str:
    return str(name or "").strip().lower().replace(" ", "_").replace("(", "").replace(")", "")


def iter_files(root: Path, pattern: str, recursive: bool) -> Iterable[Path]:
    yield from (root.rglob(pattern) if recursive else root.glob(pattern))


def read_image_shape(path: Path) -> tuple[int, int]:
    with Image.open(path) as im:
        arr = np.asarray(im)
    if arr.ndim < 2:
        raise ValueError(f"Image has no 2D shape: {path}")
    return int(arr.shape[0]), int(arr.shape[1])


def find_image(image_dir: Path, image_name: str, recursive: bool) -> Path:
    direct = image_dir / image_name
    if direct.exists():
        return direct
    stem = Path(image_name).stem
    candidates = []
    for ext in IMG_EXTS:
        candidates.extend(iter_files(image_dir, stem + ext, recursive))
    if not candidates:
        raise FileNotFoundError(f"No original image found for QuantGUV image name {image_name!r}")
    return sorted(candidates)[0]


def get_value(row: dict[str, object], aliases: Sequence[str]) -> object | None:
    for alias in aliases:
        key = norm_col(alias)
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def render_disk(mask: np.ndarray, xc: float, yc: float, radius: float, value: int) -> None:
    h, w = mask.shape
    radius = max(0.0, float(radius))
    x0 = max(0, int(np.floor(xc - radius)))
    x1 = min(w, int(np.ceil(xc + radius + 1)))
    y0 = max(0, int(np.floor(yc - radius)))
    y1 = min(h, int(np.ceil(yc + radius + 1)))
    if x0 >= x1 or y0 >= y1:
        return
    yy, xx = np.ogrid[y0:y1, x0:x1]
    region = (xx - float(xc)) ** 2 + (yy - float(yc)) ** 2 <= radius**2
    sub = mask[y0:y1, x0:x1]
    sub[region] = value


def load_rows(excel_path: Path, sheet_name: str | None) -> list[dict[str, object]]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        return []
    headers = [norm_col(value) for value in raw_rows[0]]
    rows = []
    for values in raw_rows[1:]:
        if not any(value is not None for value in values):
            continue
        rows.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
    return rows


def radius_pixels(row: dict[str, object], scale_factor: float | None) -> float:
    diameter = get_value(row, ("diameter_pixels", "diameter_px", "diameter_pix"))
    if diameter is not None:
        return float(diameter) / 2.0
    diameter_um = get_value(row, ("diameter_um", "diameter_µm", "diameter_Âµm"))
    if diameter_um is not None:
        if not scale_factor or scale_factor <= 0:
            raise ValueError("QuantGUV diameter is in microns; pass --scale-factor microns_per_pixel")
        return float(diameter_um) / float(scale_factor) / 2.0
    area = get_value(row, ("area_pixels", "area_pix", "area_px"))
    if area is not None:
        return math.sqrt(float(area) / math.pi)
    raise ValueError("Row lacks diameter or area column")


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    excel_path = Path(args.excel)
    image_dir = Path(args.image_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    rows = load_rows(excel_path, args.sheet)
    grouped: dict[str, list[dict[str, object]]] = {}
    for row in rows:
        image_name = get_value(row, ("image_name", "image", "filename", "file_name"))
        if image_name is None:
            raise ValueError("QuantGUV export lacks an Image Name column")
        grouped.setdefault(str(image_name), []).append(row)

    manifest_rows = []
    for image_name, image_rows in sorted(grouped.items()):
        image_path = find_image(image_dir, image_name, args.recursive)
        output_path = output_dir / f"{Path(image_name).stem}_pred_masks.npy"
        if output_path.exists() and not args.overwrite:
            raise FileExistsError(f"Refusing to overwrite {output_path}; pass --overwrite")
        mask = np.zeros(read_image_shape(image_path), dtype=np.int32)
        for out_id, row in enumerate(image_rows, start=1):
            x = get_value(row, ("x_pix", "x_px", "x", "center_x", "center_x_pix", "center_x_px"))
            y = get_value(row, ("y_pix", "y_px", "y", "center_y", "center_y_pix", "center_y_px"))
            if x is None or y is None:
                raise ValueError(
                    "QuantGUV export lacks center coordinates. Use the patched GUI in paper/QuantGUV "
                    "or add columns named Center X (pix) and Center Y (pix)."
                )
            render_disk(mask, float(x), float(y), radius_pixels(row, args.scale_factor), out_id)
        np.save(output_path, mask.astype(np.int32, copy=False))
        manifest_rows.append({
            "excel": str(excel_path),
            "image": str(image_path),
            "output": str(output_path),
            "n_instances": int(mask.max()),
        })

    manifest_path = output_dir / "quantguv_adapter_manifest.csv"
    with manifest_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["excel", "image", "output", "n_instances"])
        writer.writeheader()
        writer.writerows(manifest_rows)
    (output_dir / "quantguv_adapter_summary.json").write_text(json.dumps({"n_files": len(manifest_rows)}, indent=2))
    print(f"converted QuantGUV detections for {len(manifest_rows)} images")
    print(f"wrote manifest: {manifest_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
